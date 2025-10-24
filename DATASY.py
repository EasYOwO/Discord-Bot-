import discord
from discord import app_commands
from discord.ext import commands
import openpyxl
import os
import subprocess

# Intents setup
intents = discord.Intents.default()
intents.message_content = True
intents.guilds = True
intents.members = True

# Define the bot with intents and command prefix
bot = commands.Bot(command_prefix='!', intents=intents)

# Replace this with your bot token
TOKEN = ''

@bot.event
async def on_ready():
    await bot.tree.sync()  # Sync the slash commands with Discord
    print(f'Bot is online as {bot.user}')

@bot.tree.command(name="createfile")
async def create_excel_file(interaction: discord.Interaction, file_name: str):
    """Creates a new Excel file with five columns and opens it."""
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Column 1", "Column 2", "Column 3", "Column 4", "Column 5"]  # Five columns
        sheet.append(headers)
        workbook.save(file_name)

        # Open the file depending on the operating system
        if os.name == 'nt':  # Windows
            os.startfile(file_name)
        else:  # macOS or Linux
            try:
                subprocess.run(['open', file_name], check=True)  # macOS
            except FileNotFoundError:
                subprocess.run(['xdg-open', file_name], check=True)  # Linux

        await interaction.response.send_message(f"Excel file '{file_name}' created and opened successfully.", ephemeral=True)
    except Exception as e:
        await interaction.response.send_message(f"Error creating or opening Excel file: {e}", ephemeral=True)

@bot.tree.command(name="insertdata")
async def insert_data(interaction: discord.Interaction, file_name: str, column1: str, column2: str, column3: str, column4: str, column5: str):
    """Inserts a row of data into the specified Excel file."""
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        sheet.append([column1, column2, column3, column4, column5])
        workbook.save(file_name)
        await interaction.response.send_message(f"Data inserted into '{file_name}' successfully.", ephemeral=True)
    except Exception as e:
        await interaction.response.send_message(f"Error inserting data into Excel file: {e}", ephemeral=True)

@bot.tree.command(name="readdata")
async def read_data(interaction: discord.Interaction, file_name: str):
    """Reads data from the specified Excel file, saves it to a .txt file with table formatting, and sends it as an attachment."""
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            await interaction.response.send_message("No data found in the Excel file.", ephemeral=True)
            return

        # Determine column widths with added padding
        extra_padding = 3  # Increase padding to ensure enough space
        column_widths = [max(len(str(cell)) for cell in column) + extra_padding for column in zip(*rows)]

        # Create a .txt file to save the data with table formatting using 'utf-8' encoding
        txt_file_name = file_name.replace('.xlsx', '.txt')
        with open(txt_file_name, 'w', encoding='utf-8') as txt_file:
            for row in rows:
                formatted_row = " | ".join(f"{str(cell).ljust(column_widths[i])}" for i, cell in enumerate(row))
                txt_file.write(formatted_row + "\n")

        # Send the .txt file as an attachment
        with open(txt_file_name, 'rb') as file:
            await interaction.response.send_message(file=discord.File(file, txt_file_name))

        # Optionally, remove the .txt file after sending
        os.remove(txt_file_name)

    except Exception as e:
        await interaction.response.send_message(f"Error reading data from Excel file: {e}", ephemeral=True)

@bot.tree.command(name="deletefile")
async def delete_file(interaction: discord.Interaction, file_name: str):
    """Deletes a specific file from the bot's storage."""
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    if os.path.isfile(file_name):
        try:
            os.remove(file_name)
            await interaction.response.send_message(f"File '{file_name}' deleted successfully.", ephemeral=True)
        except Exception as e:
            await interaction.response.send_message(f"Error deleting file: {e}", ephemeral=True)
    else:
        await interaction.response.send_message(f"File '{file_name}' does not exist.", ephemeral=True)

@bot.tree.command(name="renamefile")
async def rename_file(interaction: discord.Interaction, old_name: str, new_name: str):
    """Renames a specific file from the bot's storage."""
    if not old_name.endswith('.xlsx'):
        old_name += '.xlsx'
    if not new_name.endswith('.xlsx'):
        new_name += '.xlsx'

    if os.path.isfile(old_name):
        try:
            os.rename(old_name, new_name)
            await interaction.response.send_message(f"File renamed from '{old_name}' to '{new_name}' successfully.", ephemeral=True)
        except Exception as e:
            await interaction.response.send_message(f"Error renaming file: {e}", ephemeral=True)
    else:
        await interaction.response.send_message(f"File '{old_name}' does not exist.", ephemeral=True)

# Run the bot
bot.run(TOKEN)

